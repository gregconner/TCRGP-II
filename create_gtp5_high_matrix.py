import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Clean, consistent, human-readable dataset
MATRIX_ROWS = [
    ["Category", "Metric", "E' Numu Diip (Land)", "River Select Foods (Fisheries)", "Many Nations (Insurance)", "RTZ Artists (Gallery)"],

    ["Membership", "Current active members", "17", "5 core members", "250 members", "~40 adults (+ youth)"],
    ["Membership", "Membership growth trend", "Stable", "Growing (clients expanding)", "Expanding", "Declined ~33% post‑COVID"],
    ["Membership", "Member retention", "High among landowners", "Stable core founders", "High", "Impacted by COVID & burnout"],

    ["Organization", "Year established (co-op)", "2017", "2014", "2009 (co-op; org 1990)", "2019"],
    ["Organization", "Years from concept to formalization", "~3 years", "~10 years", "~19 years (to co-op)", "<1 year (rapid launch)"],
    ["Organization", "Board & governance", "Board of ~5", "5 founding members", "Indigenous board", "Volunteer board"],
    ["Organization", "Key-person dependency risk", "High (project manager)", "Moderate", "Low", "High (2 people)"],

    ["Employment", "Full-time employees", "0", "2", "10", "0"],
    ["Employment", "Part-time / contract staff", "1 consultant", "2", "20 advisors", "2 contractors"],
    ["Employment", "Total jobs created (approx.)", "1", "4", "30+", "2 (plus volunteers)"],

    ["Financial", "Annual revenue (current)", "$25,000–$50,000", "$250,000–$1,000,000", ">$200,000", "$25,000–$50,000"],
    ["Financial", "Primary revenue streams", "2 (grants, land leasing)", "3 (logistics, production, consulting)", "2 (brokerage, consulting)", "1 (gallery sales)"],
    ["Financial", "Financial review cadence", "Monthly", "Quarterly", "Monthly + Quarterly board", "Ad hoc (needs improvement)"],

    ["Partnerships", "Active external partners", "~6 orgs", "170 First Nations & clients", "170 First Nations", "~5 orgs"],
    ["Partnerships", "Non‑Indigenous partners", "~3", "Processors / logistics", "Insurance carriers / vendors", "~3"],
    ["Partnerships", "Capital & funding approach", "5 grants leveraged", "Self‑financing + line of credit", "Debt‑free; retained surpluses", "2 grants + donations"],

    ["Innovation", "Major innovations (count)", "4 (solar, organic, grazing, water)", "5 (traceability, value‑add, branding, QR, engineered COGS)", "4 (wellness account, custom coverage, education fund)", "3 (apprenticeships, consignment, artist networking)"],
    ["Innovation", "New products/services", "Seed farm (planned)", "6+ branded products", "Culturally‑tailored benefits", "Gallery + training + demos"],
    ["Innovation", "Training & capacity", "Planning; need staffing", "Community workshops; model tools", "Governance & education funds", "3 apprenticeships/year; workshops"],
    ["Innovation", "Community engagement", "Limited to date", "30+ sessions over years", "Regional meetings; surveys", "4+ local events; demos"],

    ["Operations", "Production / footprint scope", "640 acres (single site)", "Distributed (national)", "National customer base", "Single gallery, on‑territory"],
    ["Operations", "Market reach", "Local / regional", "National (Canada)", "National (Canada)", "Local community"],
    ["Operations", "Primary sales channels", "Direct (planned / pilot)", "Mixed wholesale + retail", "B2B brokerage", "Direct gallery"],
    ["Operations", "Digital tools used", "2 (website, Facebook)", "4 (web, QR, profiles, social)", "3 (web, social, systems)", "2 (web, social)"],

    ["Culture", "Cultural values integrated", "3 (food, language, land)", "4 (stories, ethics, sovereignty, place)", "5 (wellness, customs, community)", "4 (tradition, community, customs)"],
    ["Culture", "Preservation initiatives", "Traditional foods & language", "Conservation & food security", "Traditional wellness coverage", "Authentic art; knowledge sharing"],
    ["Culture", "Multi‑tribal involvement", "No (single‑tribe focus)", "Yes (170 nations)", "Yes (170 nations)", "No (single‑tribe focus)"],

    ["Challenges", "Major persistent challenges", "5 (BIA, water, funding, capacity, politics)", "6 (collapse, brokers, resistance, inventory)", "4 (barriers, engagement, transparency)", "6 (trust, COVID, engagement, burnout)"],
    ["Challenges", "Gov't / regulatory issues", "Yes – BIA", "Yes – DFO", "Regulatory complexity", "Minimal"],
    ["Challenges", "Community relations", "Some tensions", "Broker conflicts legacy", "Strong/member‑led", "Trust rebuild needed"],
    ["Challenges", "Risk mitigation", "3 (diversify, partners, planning)", "4 (value‑add, diversify, inventory, partners)", "Scale & transparency", "Limited formalization"],
    ["Challenges", "Climate/market adaptation", "3 (regenerative, conservation, traditional)", "4 (sustainable, selective, local value‑add)", "Market education", "COVID pivots"],

    ["Strategy", "Unique value proposition", "Land sovereignty & sustainability", "Indigenous‑branded seafood w/ traceability", "Culturally‑appropriate insurance", "Artist‑controlled pricing & mentorship"],
    ["Strategy", "Competitive advantage", "Owns 640 acres; autonomy", "Vertical integration & brand", "Scale + expertise + trust", "Authenticity + community ties"],
    ["Strategy", "Growth potential (2‑3 yrs)", "High (seed farm; solar)", "High (channel expansion)", "Moderate (saturation risk)", "High (if trust rebuilt)"],
    ["Strategy", "Sustainability outlook", "Strong (asset‑backed)", "Strong (diversified)", "Excellent (profitable)", "Vulnerable (volunteer‑reliant)"],
]


def save_csv(path: str):
    with open(path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(["Indigenous Cooperative Comparative Analysis – gtp-5-high v1.0.0"])
        writer.writerow([])
        for row in MATRIX_ROWS:
            writer.writerow(row)


def save_xlsx(path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Analysis"

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = "Indigenous Cooperative Comparative Analysis – gtp-5-high v1.0.0"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1A252F", end_color="1A252F", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Empty spacer
    # Header row at row 3
    for r, row in enumerate(MATRIX_ROWS, start=3):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)

    # Styles
    header_row = 3
    for c in range(1, 7):
        cell = ws.cell(row=header_row, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Section fills by Category
    fills = {
        'Membership': "E8F4FD",
        'Organization': "F3E5F5",
        'Employment': "E8F5E9",
        'Financial': "FFF8E1",
        'Partnerships': "FFF3E0",
        'Innovation': "FCE4EC",
        'Operations': "E0F7FA",
        'Culture': "FFF9C4",
        'Challenges': "FFEBEE",
        'Strategy': "EDE7F6",
    }

    thin = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
                  top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))

    for r in range(4, ws.max_row + 1):
        cat = ws.cell(row=r, column=1).value
        color = fills.get(cat, "FFFFFF")
        for c in range(1, 7):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.border = thin
            if c == 1:
                cell.font = Font(bold=True)
            elif c == 2:
                cell.font = Font(italic=True)
            else:
                val = str(cell.value)
                if any(x in val for x in ["High", "Strong", "Excellent", "170", "30+"]):
                    cell.font = Font(color="006600", bold=True)
                elif any(x in val for x in ["Vulnerable", "Declined", "0 (", "0$", "Minimal"]):
                    cell.font = Font(color="CC0000")

    # Column widths
    widths = [16, 36, 24, 24, 24, 24]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64+i)].width = w

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[3].height = 24

    wb.save(path)


if __name__ == '__main__':
    save_csv('gtp-5-high_beautiful_cooperative_matrix_v1.0.0.csv')
    save_xlsx('gtp-5-high_beautiful_cooperative_matrix_v1.0.0.xlsx')
    print('Matrix generated: gtp-5-high v1.0.0 (CSV/XLSX)')
