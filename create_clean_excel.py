import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Read the clean CSV data
with open('claude-sonnet-4-max_clean_readable_matrix_v1.3.0.csv', 'r', encoding='utf-8') as f:
    reader = csv.reader(f)
    data = list(reader)

# Create new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Cooperative Analysis"

# Write data to worksheet
for row_idx, row_data in enumerate(data, 1):
    for col_idx, cell_value in enumerate(row_data, 1):
        ws.cell(row=row_idx, column=col_idx, value=cell_value)

# Apply formatting - clean professional style like Sonnet version
header_font = Font(name='Calibri', size=11, bold=True)
data_font = Font(name='Calibri', size=11)
category_font = Font(name='Calibri', size=11, bold=True)

# Header row formatting
for col in range(1, len(data[0]) + 1):
    cell = ws.cell(row=1, column=col)
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Category column formatting (column A)
for row in range(2, len(data) + 1):
    cell = ws.cell(row=row, column=1)
    if cell.value and cell.value.strip():  # If category name exists
        cell.font = category_font
        cell.alignment = Alignment(horizontal='left', vertical='center')
    else:
        cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)

# Data formatting for all other columns
for row in range(2, len(data) + 1):
    for col in range(2, len(data[0]) + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = data_font
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Set column widths for readability
ws.column_dimensions['A'].width = 25  # Category
ws.column_dimensions['B'].width = 45  # Question
ws.column_dimensions['C'].width = 35  # Legend
ws.column_dimensions['D'].width = 20  # E' Numu
ws.column_dimensions['E'].width = 20  # RSF
ws.column_dimensions['F'].width = 20  # Many Nations
ws.column_dimensions['G'].width = 20  # RTZ

# Set row heights
for row in range(1, len(data) + 1):
    ws.row_dimensions[row].height = 20

# Freeze panes at header row
ws.freeze_panes = 'A2'

# Save the Excel file
wb.save('claude-sonnet-4-max_clean_readable_matrix_v1.3.0.xlsx')

print('✅ Created clean Excel file!')
print('📊 claude-sonnet-4-max_clean_readable_matrix_v1.3.0.xlsx')
print('')
print('🎨 Formatting applied:')
print('  • Clean professional fonts (Calibri)')
print('  • No distracting colors')
print('  • Proper column widths for readability')
print('  • Category headers left-aligned')
print('  • Data properly wrapped and aligned')
print('  • Frozen header row for easy scrolling')

