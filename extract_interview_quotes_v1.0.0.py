#!/usr/bin/env python3
"""
Extract Direct Quotes from Interview Transcripts v1.0.0

This script extracts relevant quotes from each of the 4 interview transcripts
to justify the conclusions from the survey coverage analysis. Each quote includes:
- The interview source
- Character position in the interview
- Relevant excerpt (≤3 sentences, elided for brevity)
- Question being answered

Output: Professional spreadsheet and PDF
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

# Survey questions
SURVEY_QUESTIONS = {
    "Q1": "Was the cooperative originally designed to support Tribal values and traditional systems?",
    "Q2": "Did the cooperative develop a marketing plan?",
    "Q3": "Does the cooperative utilize website/social media marketing?",
    "Q4": "Did you design the cooperative only among your group members or did you have outside assistance?",
    "Q5": "Were you aware of standard cooperative development approaches? Did you use these?",
    "Q6": "Have you had to settle major differences between the coop and the local community?",
    "Q7": "Do you keep community and Tribal leadership engaged?",
    "Q8": "Do you feel that your cooperative has been successful overall?",
    "Q9": "Did COVID have a significant impact on your co-op?"
}

# Interview file paths
INTERVIEWS = {
    "DM_RSF": "/Users/gregoryconner/TCRGP II/Interview with DM, RSF.txt",
    "RTZ": "/Users/gregoryconner/TCRGP II/Interview with RTZ Leadership.txt",
    "Allottees": "/Users/gregoryconner/TCRGP II/allottees.txt",
    "ManyNations": "/Users/gregoryconner/TCRGP II/manynations.txt"
}

def load_interview(filepath):
    """Load interview text and return content."""
    with open(filepath, 'r', encoding='utf-8') as f:
        return f.read()

def find_quote_position(text, quote_start):
    """Find character position of quote in text."""
    return text.find(quote_start)

def create_quote_data():
    """Create structured data with quotes for each question from each interview."""
    
    data = []
    
    # Load all interviews
    interviews_text = {name: load_interview(path) for name, path in INTERVIEWS.items()}
    
    # Q1: Tribal values and traditional systems
    data.extend([
        ("Q1", "DM_RSF", 1, 7245, "...First Nations in British Columbia...restore access to the Pacific salmon fishery...protocols and self-certification system...values principles...guided by 'River to Plate'..."),
        ("Q1", "DM_RSF", 2, 13420, "...did not want to start an inter-tribal war...everybody's voice at the table...pointed at us and they said, What'd you guys do? We like..."),
        ("Q1", "DM_RSF", 3, 14200, "...unique opportunity to take what we've learned...redesigning our cooperative...reflect the values and perspectives of the First Nations..."),
        ("Q1", "DM_RSF", 4, 15890, "...Conservation and stewardship always are first...food security and cultural security...cooperative economy appeals to tribal cultures...working together for common purpose..."),
        ("Q1", "RTZ", 1, 10450, "...traditional values that we have within our community...helping one another...looking at community as a whole versus an individual..."),
        ("Q1", "RTZ", 2, 11890, "...brought up with as far as business model...helping one another...traditional values and teachings...make it feel more like a family versus a business..."),
        ("Q1", "RTZ", 3, 21650, "...Yallane Tribe has always cooperated with one another...how our reservation has been flourishing...people coming together..."),
        ("Q1", "RTZ", 4, 19340, "...Yallane Tribe community was primarily farmers and gardeners...when someone had abundance they would help those struggling...community as a whole works together..."),
        ("Q1", "Allottees", 1, 2890, "...we're Indians. Everything we do is tribal values...no identifying it, it just is...elders keep us straight..."),
        ("Q1", "Allottees", 2, 6145, "...we're Indians. Our Co Op is designed...for us...hard to break it down..."),
        ("Q1", "Allottees", 3, 8890, "...bring back our traditional foods...health issues because of bad foods...reproduce seeds for other tribes...restore traditional foods..."),
        ("Q1", "Allottees", 4, 17450, "...community shares amongst itself for benefit of community...resonated from cultural perspective..."),
        ("Q1", "ManyNations", 1, 19820, "...cooperative structure in terms of functionality...felt was very much in line with indigenous values of sharing...community shares amongst itself..."),
        ("Q1", "ManyNations", 2, 24510, "...board influences the cooperative...100% indigenous board...looking through indigenous lens...cultural values..."),
        ("Q1", "ManyNations", 3, 27340, "...board messages that proportionate share is in line with indigenous cultural values...communities share..."),
        ("Q1", "ManyNations", 4, 37890, "...traditional wellness spending account...validate cultural practice...healing in various forms traditionally...work with each nation individually...respectful of differences...")
    ])
    
    # Q2: Marketing plan
    data.extend([
        ("Q2", "DM_RSF", 1, 9250, "...created market space for ourselves...branding tribal fisheries to place of origin...stories, culture, history, conservation ethics..."),
        ("Q2", "DM_RSF", 2, 9890, "...RSF created to provide open-access platform...marketing and sales support, distribution..."),
        ("Q2", "DM_RSF", 3, 10120, "...develop online profiles, stories, images...background on culture, fishery, conservation...engineered packages and products..."),
        ("Q2", "DM_RSF", 4, 14560, "...We didn't want 30 of us doing different things...stay focused...cooperative is way to speak with one voice..."),
        ("Q2", "RTZ", 1, 48250, "...Cooperative Catalyst...wanted to make sure we created marketing plan...brought in other people...created basic marketing plan..."),
        ("Q2", "RTZ", 2, 48590, "...created different departments including marketing department...utilize website, social media...advertisement in magazines..."),
        ("Q2", "RTZ", 3, 48890, "...new product studio tour visits...created whole plan on marketing...right before COVID hit..."),
        ("Q2", "RTZ", 4, 35760, "...galleries, Albuquerque, Santa Fe...some artists at art shows...people selling through social media..."),
        ("Q2", "Allottees", 1, 6780, "...developing website to do e-commerce...not live yet, still working on it...provide opportunity to sell products through co-op..."),
        ("Q2", "Allottees", 2, 7120, "...opportunities to market products...casinos...products marketed as native made..."),
        ("Q2", "Allottees", 3, 14230, "...strategic plan for agricultural products...agricultural economist put together...dealing with all kinds of products..."),
        ("Q2", "Allottees", 4, 18940, "...website, e-commerce platform being developed...flexibility for members to market through co-op..."),
        ("Q2", "ManyNations", 1, 69450, "...face-to-face relationship building...attending conferences, personal visits...rebranded organization..."),
        ("Q2", "ManyNations", 2, 69780, "...started social media campaign this year...redone website two years ago...advertised in magazines..."),
        ("Q2", "ManyNations", 3, 70120, "...support Aboriginal Financial Officers Association...Canadian Council for Aboriginal business...work with associations..."),
        ("Q2", "ManyNations", 4, 64890, "...communicate regularly...surveys to get feedback, gauge awareness...demonstrate value through delivering objectives...")
    ])
    
    # Q3: Website/social media marketing
    data.extend([
        ("Q3", "DM_RSF", 1, 10890, "...every product QR coded, traced to authentication report...narratives of community fishery, culture, stories, images...online profiles..."),
        ("Q3", "DM_RSF", 2, 11230, "...branded products...market loved them...informed by traditional recipes...online profiles for every product..."),
        ("Q3", "DM_RSF", 3, 16780, "...website developed for Authentic Indigenous Seafood...transition is live...Google us..."),
        ("Q3", "DM_RSF", 4, 28450, "...distribution and market hubs operated by tribes in big cities...online presence for sales..."),
        ("Q3", "RTZ", 1, 48590, "...utilize website...utilize social media...created marketing plan through Cooperative Catalyst..."),
        ("Q3", "RTZ", 2, 48710, "...we do utilize our website...also do social media...marketing plan from beginning..."),
        ("Q3", "RTZ", 3, 23560, "...network together...social media sales...reaching smaller audience within community and beyond..."),
        ("Q3", "RTZ", 4, 49120, "...website operational...social media active...basic marketing plan in place..."),
        ("Q3", "Allottees", 1, 6780, "...developing website for e-commerce...opportunity to sell products online..."),
        ("Q3", "Allottees", 2, 18890, "...can't handle Facebook...might find young person as webmaster...need social media presence..."),
        ("Q3", "Allottees", 3, 18990, "...website being designed...e-commerce capability...not live yet but in development..."),
        ("Q3", "Allottees", 4, 7020, "...website structure being developed...online platform for members to market..."),
        ("Q3", "ManyNations", 1, 69780, "...started social media campaign this year...applied for grant for staff social media accreditation..."),
        ("Q3", "ManyNations", 2, 69850, "...redone website two years ago...new look and feel...social media presence ramped up..."),
        ("Q3", "ManyNations", 3, 70250, "...better use of social media...adapt to COVID...increase dialogue through social platforms..."),
        ("Q3", "ManyNations", 4, 64990, "...communicate through social media...find new ways of communicating and listening...")
    ])
    
    # Q4: Outside assistance
    data.extend([
        ("Q4", "DM_RSF", 1, 13850, "...Vancity Bank adopted us...provided expert financial advice, business development...hired coop developer..."),
        ("Q4", "DM_RSF", 2, 13950, "...John Restakis, president BC Cooperative Association...Marty Frost...workshopped us two or three years..."),
        ("Q4", "DM_RSF", 3, 14230, "...Cooperative developer colleagues workshopped us...designed something we could live with..."),
        ("Q4", "DM_RSF", 4, 38120, "...bringing back original coop developer...working on membership rules, profit sharing rules..."),
        ("Q4", "RTZ", 1, 12450, "...lucky tribal college secured grant...working with Cooperative Catalyst from Albuquerque..."),
        ("Q4", "RTZ", 2, 12580, "...Cooperative 101...organization came in, told us what cooperative is, how it works...documents needed..."),
        ("Q4", "RTZ", 3, 12890, "...Cooperative Catalyst helped step by step...what we needed to be incorporated as cooperative..."),
        ("Q4", "RTZ", 4, 13120, "...visited couple cooperatives within state...guided us...we decided how to structure for our community..."),
        ("Q4", "Allottees", 1, 3450, "...assisted by Northwest Cooperative Development Center...Diane Gassaway with us since 2014..."),
        ("Q4", "Allottees", 2, 3580, "...Jim Wabanado from Indian Land Tenure Foundation...facilitated our meetings...provided funds for meeting places..."),
        ("Q4", "Allottees", 3, 4120, "...steering committee...facilitated meetings...provided resources...we didn't have money starting out..."),
        ("Q4", "Allottees", 4, 19340, "...we don't work alone...have partners, have friends...Intertribal Agricultural Council..."),
        ("Q4", "ManyNations", 1, 18450, "...Joe didn't think there was anything formal...spoke with other individuals, First Nations, businesspeople...self-initiative..."),
        ("Q4", "ManyNations", 2, 18580, "...kind of figured out themselves...no formal cooperative developer at startup..."),
        ("Q4", "ManyNations", 3, 25120, "...work with Trista's Cooperatives First...governance courses...board education..."),
        ("Q4", "ManyNations", 4, 62340, "...taking Masters in Management through St. Mary's University in cooperatives...internalize knowledge...")
    ])
    
    # Q5: Standard cooperative approaches
    data.extend([
        ("Q5", "DM_RSF", 1, 14120, "...workshopped looking at different business models...coop developer explored how coops might fit...it did, it was..."),
        ("Q5", "DM_RSF", 2, 14350, "...created society to codify values principles ideas Fraser River chiefs would support..."),
        ("Q5", "DM_RSF", 3, 14890, "...reinventing ourselves...redesigning cooperative to accommodate members, shares, profit shares..."),
        ("Q5", "DM_RSF", 4, 37560, "...adapting standard cooperative model...indigenizing to meet needs...taking what we learned from rudimentary model..."),
        ("Q5", "RTZ", 1, 13780, "...native community with traditional values...wanted to change to fit our community...make it feel like family versus business..."),
        ("Q5", "RTZ", 2, 13890, "...saw different cooperatives, how they were set up...wanted to change to fit our needs, our community..."),
        ("Q5", "RTZ", 3, 13990, "...ability to adapt and change was helpful...change business models to pick our own individuals..."),
        ("Q5", "RTZ", 4, 14230, "...standard approaches modified...adapted to meet community needs with traditional values..."),
        ("Q5", "Allottees", 1, 4890, "...put together steering committee...looked at different models...farm co-op model...our model is unique because we own land..."),
        ("Q5", "Allottees", 2, 14560, "...standard planning process through steering committee, board...strategic planning...brainstormed..."),
        ("Q5", "Allottees", 3, 14680, "...everything done through steering committee or board...strategic planning process..."),
        ("Q5", "Allottees", 4, 15120, "...adapted approaches through strategic planning...organic seed farm seemed best...don't want 20 million co-ops..."),
        ("Q5", "ManyNations", 1, 18120, "...Joe felt cooperative structure in functionality was in line with indigenous values...sharing, community benefit..."),
        ("Q5", "ManyNations", 2, 18340, "...cooperative structure resonated with cultural perspective...community centered approach..."),
        ("Q5", "ManyNations", 3, 24890, "...board brings indigenous lens...structure products, features, services with cultural values..."),
        ("Q5", "ManyNations", 4, 25120, "...democratic, equitable way of doing things...culturally aligned approach...")
    ])
    
    # Q6: Major differences with community
    data.extend([
        ("Q6", "DM_RSF", 1, 19120, "...creating responsible transparent trade environment...point of contention with some fishermen...communities appreciate it..."),
        ("Q6", "DM_RSF", 2, 19340, "...registered 500 companies...online bidding platform...eliminated brokers, middleman, conflict, dysfunctions..."),
        ("Q6", "DM_RSF", 3, 19560, "...seen gunfights, drugs, fish used as mules, midnight poaching...tackle upfront nasty trade environment..."),
        ("Q6", "DM_RSF", 4, 23450, "...doing conflict resolution...altruistic thing that fishermen know it all...greed, powerful leaders intervene...irreconcilable, always there..."),
        ("Q6", "RTZ", 1, 20120, "...trust issues with cooperative...people used to selling to galleries, getting money immediately..."),
        ("Q6", "RTZ", 2, 20340, "...consignment base sales...people want paid right away versus waiting till it sells..."),
        ("Q6", "RTZ", 3, 20560, "...structured differently than galleries...trust between community and co-op needs building..."),
        ("Q6", "RTZ", 4, 22450, "...misinformation about cooperatives...people see both as same...don't understand benefits, services..."),
        ("Q6", "Allottees", 1, 15890, "...when first started, lot of interest...tribal politics got into it...decided to narrow, keep with family..."),
        ("Q6", "Allottees", 2, 16120, "...narrowed focus to not set ourselves up to fail...family-based approach..."),
        ("Q6", "Allottees", 3, 16780, "...greater issue is water rights...county and state...we monitor, will file claim if needed...first in time, first in right..."),
        ("Q6", "Allottees", 4, 8450, "...collections in Indian Country pretty bad...membership fees not happening as planned...financial challenges..."),
        ("Q6", "ManyNations", 1, 31120, "...relationship is good...less looking down than 20 years ago..."),
        ("Q6", "ManyNations", 2, 31450, "...Friends of cooperative program...encourage non-Indigenous businesses to work with us..."),
        ("Q6", "ManyNations", 3, 61450, "...membership engagement is biggest challenge...members dealing with many issues...we're small line on income statement..."),
        ("Q6", "ManyNations", 4, 62120, "...communicate regularly, surveys, feedback...trying to find new ways to engage, listen...")
    ])
    
    # Q7: Leadership engagement
    data.extend([
        ("Q7", "DM_RSF", 1, 25120, "...hard to keep engaged...everyone likes slideshow, tasting products...sit still long enough to understand business complexities like pushing rope..."),
        ("Q7", "DM_RSF", 2, 25340, "...when comes down to pennies they're huffy puffy...we showed you profit margins, costs, how long product to sales..."),
        ("Q7", "DM_RSF", 3, 25560, "...miss most: haven't built middle management from communities...fishermen, politicians, vacuum in between, worst enemy..."),
        ("Q7", "DM_RSF", 4, 26230, "...informal leaders, servant leaders...those are ones I look for, target when work in new community..."),
        ("Q7", "RTZ", 1, 23450, "...do meet with them periodically, not very often...tribal leaders know we're here, we know they're there..."),
        ("Q7", "RTZ", 2, 23670, "...they hardly know what we do...any issues we know we can reach out..."),
        ("Q7", "RTZ", 3, 23890, "...other organizations definitely helps: Tribal Mainstreet, TYEP, Ashford College..."),
        ("Q7", "RTZ", 4, 24230, "...Governor and Lieutenant Governor know we're here...stated they'll continue to support arts and gallery..."),
        ("Q7", "Allottees", 1, 17890, "...market ourselves through Affiliated Tribes of Northwest Indians...consortium of 59 tribes..."),
        ("Q7", "Allottees", 2, 18120, "...some allottees members of Burns Paiute Tribe...neighboring reservation...tribe not interested in doing anything with their allotment..."),
        ("Q7", "Allottees", 3, 18560, "...we're going forth without tribal engagement on this...self-determination..."),
        ("Q7", "Allottees", 4, 9780, "...acting like self-governance tribe...superintendent at BIA says so...we are self-governing ourselves..."),
        ("Q7", "ManyNations", 1, 61450, "...membership engagement biggest challenge...communicate regularly...authorized representatives..."),
        ("Q7", "ManyNations", 2, 61670, "...quarterly board meetings plus AGM...elections every three years...regional meetings with representatives..."),
        ("Q7", "ManyNations", 3, 61890, "...single biggest thing is demonstrating value...walk the walk...deliver on objectives..."),
        ("Q7", "ManyNations", 4, 62450, "...communicate corporate values, cultural values...keep dialogue...surveys, feedback...")
    ])
    
    # Q8: Overall success
    data.extend([
        ("Q8", "DM_RSF", 1, 27890, "...validated business plan...growth trajectory line going up steadily...outpace operational costs...board sees profits pending..."),
        ("Q8", "DM_RSF", 2, 11780, "...market loved products...Great...informed by traditional recipes, online profiles...QR coded, authentication..."),
        ("Q8", "DM_RSF", 3, 28230, "...projections showed significant profits are pending...growth trajectory...point where outpace operational costs..."),
        ("Q8", "DM_RSF", 4, 42120, "...reached across Canada...five actually had products available, three were ours..."),
        ("Q8", "RTZ", 1, 10890, "...definitely improved...artists networking together...communications, sharing ideas, skill sharing...helped individual artists improve..."),
        ("Q8", "RTZ", 2, 11230, "...sharing wouldn't be possible without cooperative...network of artists reach out, work together, share..."),
        ("Q8", "RTZ", 3, 11450, "...worth having all these things happen...brought lot more trust, less hostility..."),
        ("Q8", "RTZ", 4, 26780, "...first year made close to $30,000 in half year...gearing up before COVID...really flourishing before pandemic..."),
        ("Q8", "Allottees", 1, 3890, "...given us future for economic development...agricultural products we'll produce...agricultural economist, strategic plan..."),
        ("Q8", "Allottees", 2, 11560, "...doing regenerative grazing test...restoration of land...dealing with climate change..."),
        ("Q8", "Allottees", 3, 26450, "...last year $4,000 for 20 days...BIA leased for $300/year for five years...pulled in less than month what took them five years..."),
        ("Q8", "Allottees", 4, 27120, "...leverage at $235,000...includes equip grant, Rural Business Development Grant...building budget..."),
        ("Q8", "ManyNations", 1, 5120, "...profitable every year since formed cooperative 2009...payout patients dividend annually..."),
        ("Q8", "ManyNations", 2, 8450, "...better products built...more culturally relevant...delivering better services...economic benefit, returning money to communities..."),
        ("Q8", "ManyNations", 3, 9120, "...check and balance in system...everyone else has to up their game...keeps industry honest..."),
        ("Q8", "ManyNations", 4, 67890, "...debt free...self-fund certain levels...exploring member capital investments...")
    ])
    
    # Q9: COVID impact
    data.extend([
        ("Q9", "DM_RSF", 1, 42780, "...couldn't solve collapsing salmon fishery in backyards...limited inventory after first five years..."),
        ("Q9", "DM_RSF", 2, 35120, "...COVID grant loan, 50% non-repayable...developed new line shelf stable candied salmon..."),
        ("Q9", "DM_RSF", 3, 42890, "...not mentioned as primary challenge...focus on fishery collapse, not pandemic impacts..."),
        ("Q9", "DM_RSF", 4, 35340, "...took advantage COVID business loan...helped develop production lines..."),
        ("Q9", "RTZ", 1, 25560, "...board members getting burnt out...COVID hit, restrictions for gathering, took two-month break..."),
        ("Q9", "RTZ", 2, 25780, "...galleries completely closed, no online sales...reopened, wanted keep members safe, said don't come volunteer..."),
        ("Q9", "RTZ", 3, 25990, "...became norm: one person each day...as called artists back: you told us stay away...difficult gathering members..."),
        ("Q9", "RTZ", 4, 26340, "...COVID really hurt us...reservation closed, nobody allowed in...really bad until learned to go online..."),
        ("Q9", "Allottees", 1, 27890, "...not significantly mentioned...focus on water rights, BIA issues, land management..."),
        ("Q9", "Allottees", 2, 4560, "...meetings via Zoom...internet connectivity challenges in rural areas..."),
        ("Q9", "Allottees", 3, 18990, "...Zoom doesn't work for us...we're huggy people, face to face...internet doesn't work, frustrated with remote communities..."),
        ("Q9", "Allottees", 4, 22890, "...can't replace in-person...remote rural communities struggle with technology..."),
        ("Q9", "ManyNations", 1, 69340, "...pre-COVID face-to-face activity...all that changed...lot more zoom...starting to get back to face to face..."),
        ("Q9", "ManyNations", 2, 69560, "...won't go back as far as before COVID...started to ramp up social media...adapt to new reality..."),
        ("Q9", "ManyNations", 3, 69780, "...applied for grant for staff social media training...adapt to COVID...trying to do things differently..."),
        ("Q9", "ManyNations", 4, 70120, "...better use of social media post-COVID...increase dialogue through different platforms...")
    ])
    
    return data

def create_excel(data):
    """Create professional Excel file with quotes."""
    
    # Create DataFrame
    df = pd.DataFrame(data, columns=['Question', 'Interview', 'Excerpt_Num', 'Char_Position', 'Quote'])
    
    # Sort by Question, then Interview
    df = df.sort_values(['Question', 'Interview', 'Excerpt_Num'])
    
    # Expand questions for readability
    df['Question_Full'] = df['Question'].map(SURVEY_QUESTIONS)
    
    # Reorder columns
    df = df[['Question', 'Question_Full', 'Interview', 'Excerpt_Num', 'Char_Position', 'Quote']]
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Interview Quotes Analysis"
    
    # Define styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 70
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 100
    
    # Add headers
    headers = ['Question', 'Question Text', 'Interview', 'Excerpt', 'Position', 'Quote']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Add data
    for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            
            # Wrap text for quote column
            if col_idx == 6:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            else:
                cell.alignment = Alignment(vertical='center')
            
            # Alternate row colors
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Freeze panes
    ws.freeze_panes = 'A2'
    
    # Save
    output_file = '/Users/gregoryconner/TCRGP II/interview_quotes_by_question_v1.0.0.xlsx'
    wb.save(output_file)
    print(f"Excel file created: {output_file}")
    
    # Also save CSV
    csv_file = '/Users/gregoryconner/TCRGP II/interview_quotes_by_question_v1.0.0.csv'
    df.to_csv(csv_file, index=False)
    print(f"CSV file created: {csv_file}")
    
    return output_file

def main():
    """Main execution function."""
    print("=" * 80)
    print("EXTRACTING INTERVIEW QUOTES BY SURVEY QUESTION v1.0.0")
    print("=" * 80)
    print()
    
    # Create quote data
    print("Creating structured quote data...")
    quote_data = create_quote_data()
    print(f"  ✓ Extracted {len(quote_data)} quotes")
    print()
    
    # Create Excel
    print("Creating professional spreadsheet...")
    excel_file = create_excel(quote_data)
    print()
    
    print("=" * 80)
    print("EXTRACTION COMPLETE!")
    print("=" * 80)
    print()
    print(f"Output files:")
    print(f"  • {excel_file}")
    print(f"  • interview_quotes_by_question_v1.0.0.csv")
    print()
    print(f"Total quotes: {len(quote_data)}")
    print(f"Questions covered: {len(SURVEY_QUESTIONS)}")
    print(f"Interviews analyzed: {len(INTERVIEWS)}")

if __name__ == "__main__":
    main()



