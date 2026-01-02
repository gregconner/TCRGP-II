import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd

print('🔍 Examining the previous beautiful matrix Excel formatting...')
print()

# Load the previous beautiful matrix Excel file
wb = openpyxl.load_workbook('conner_claude-4-sonnet-1m_beautiful_cooperative_matrix_v2.0.xlsx')
ws = wb.active

print(f'📊 Worksheet name: {ws.title}')
print(f'📏 Dimensions: {ws.max_row} rows x {ws.max_column} columns')
print()

# Check the structure and formatting
print('📋 Column headers:')
for col in range(1, ws.max_column + 1):
    cell = ws[f'{openpyxl.utils.get_column_letter(col)}1']
    print(f'  Col {col}: "{cell.value}"')
    print(f'    Font: {cell.font.name}, Size: {cell.font.size}, Bold: {cell.font.bold}')
    if cell.fill.start_color.index != '00000000':
        print(f'    Fill: {cell.fill.start_color.index}')
    print(f'    Alignment: {cell.alignment.horizontal}')
print()

# Check first few data rows formatting
print('🎨 Sample row formatting (rows 2-5):')
for row in range(2, min(6, ws.max_row + 1)):
    print(f'Row {row}:')
    for col in range(1, min(4, ws.max_column + 1)):  # Just first 3 columns
        cell = ws.cell(row=row, column=col)
        print(f'  Col {col}: "{cell.value}"')
        print(f'    Font: {cell.font.name}, Size: {cell.font.size}, Bold: {cell.font.bold}')
        if hasattr(cell.fill, 'start_color') and cell.fill.start_color.index != '00000000':
            print(f'    Fill: {cell.fill.start_color.index}')
        print(f'    Alignment: {cell.alignment.horizontal}, Indent: {cell.alignment.indent}')
    print()

# Check column widths
print('📐 Column widths:')
for col in range(1, ws.max_column + 1):
    col_letter = openpyxl.utils.get_column_letter(col)
    width = ws.column_dimensions[col_letter].width
    print(f'  Column {col} ({col_letter}): {width}')

print()
print('🔍 Key formatting patterns identified!')

