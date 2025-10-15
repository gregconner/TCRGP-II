import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import csv

# Read the cited data
with open('claude-sonnet-4-max_with_citations_v1.8.0.csv', 'r', encoding='utf-8') as f:
    reader = csv.reader(f)
    data = list(reader)

# Create Excel with citations
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Cooperative Analysis with Citations"

# Fonts and colors
title_font = Font(name='Calibri', size=16, bold=True)
header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
category_font = Font(name='Calibri', size=11, bold=True)
question_font = Font(name='Calibri', size=10, bold=False)
legend_font = Font(name='Calibri', size=9, bold=False, italic=True)
data_font = Font(name='Calibri', size=10, bold=False)
citation_font = Font(name='Calibri', size=8, bold=False, color='666666')

blue_header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
light_blue_fill = PatternFill(start_color='E8F1FF', end_color='E8F1FF', fill_type='solid')
citation_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')

# Title
ws.merge_cells('A1:K1')
title_cell = ws['A1']
title_cell.value = "INDIGENOUS COOPERATIVE ANALYSIS - WITH INTERVIEW CITATIONS"
title_cell.font = title_font
title_cell.alignment = Alignment(horizontal='center', vertical='center')

# Headers
headers = ["Category", "Question", "Legend", 
          "Citation", "E' Numu Diip", "Citation", "River Select Foods", 
          "Citation", "Many Nations", "Citation", "RTZ Artists"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = blue_header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Column widths - alternating data and citation columns
widths = [20, 35, 25, 15, 12, 15, 12, 15, 12, 15, 12]
for col, width in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = width

# Data rows
current_row = 4
current_category = ""

for data_row in data[1:]:
    category = data_row[0].strip() if len(data_row) > 0 else ""
    question = data_row[1] if len(data_row) > 1 else ""
    legend = data_row[2] if len(data_row) > 2 else ""
    
    # Category header
    if category and category != current_category:
        ws.merge_cells(f'A{current_row}:K{current_row}')
        cat_cell = ws.cell(row=current_row, column=1, value=category.upper())
        cat_cell.font = category_font
        cat_cell.fill = light_blue_fill
        cat_cell.alignment = Alignment(horizontal='left', vertical='center')
        current_row += 1
        current_category = category
    
    # Question and legend
    ws.cell(row=current_row, column=2, value=question).font = question_font
    ws.cell(row=current_row, column=3, value=legend).font = legend_font
    
    # Data and citations (alternating columns)
    for i in range(4):  # 4 cooperatives
        citation_col = 4 + (i * 2)
        data_col = citation_col + 1
        
        if citation_col - 1 < len(data_row):
            citation = data_row[citation_col - 1] if citation_col - 1 < len(data_row) else ""
            value = data_row[data_col - 1] if data_col - 1 < len(data_row) else ""
            
            # Citation cell
            cite_cell = ws.cell(row=current_row, column=citation_col, value=citation)
            cite_cell.font = citation_font
            cite_cell.fill = citation_fill
            cite_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Data cell
            data_cell = ws.cell(row=current_row, column=data_col, value=value)
            data_cell.font = data_font
            data_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    current_row += 1

# Row heights
ws.row_dimensions[1].height = 25
ws.row_dimensions[3].height = 40
for row in range(4, current_row):
    ws.row_dimensions[row].height = 22

ws.freeze_panes = 'A4'

wb.save('claude-sonnet-4-max_with_citations_v1.8.0.xlsx')

print('✅ Created Excel with interview citations!')
print('📊 claude-sonnet-4-max_with_citations_v1.8.0.xlsx')
print('📄 claude-sonnet-4-max_with_citations_v1.8.0.csv')
print('')
print('📍 Features:')
print('  • Citation column next to each cooperative')
print('  • Gray background for citation columns')
print('  • Specific line numbers where available')
print('  • Inference notes for derived data')
print('  • Complete traceability to interview sources')

