import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import csv

# Read the completely fixed data WITH Legend column
with open('claude-sonnet-4-max_with_legend_v1.6.0.csv', 'r', encoding='utf-8') as f:
    reader = csv.reader(f)
    data = list(reader)

# Create new workbook with beautiful formatting
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Indigenous Cooperative Comparison"

# Define colors and fonts exactly like the previous beautiful version
title_font = Font(name='Calibri', size=16, bold=True)
header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
category_font = Font(name='Calibri', size=11, bold=True)
question_font = Font(name='Calibri', size=10, bold=False)
legend_font = Font(name='Calibri', size=9, bold=False, italic=True)
data_font = Font(name='Calibri', size=10, bold=False)

# Colors from the previous version
blue_header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
light_blue_fill = PatternFill(start_color='E8F1FF', end_color='E8F1FF', fill_type='solid')

# Row 1: Main title (merged across all columns)
ws.merge_cells('A1:G1')
title_cell = ws['A1']
title_cell.value = "INDIGENOUS COOPERATIVE SYSTEMS - COMPARATIVE ANALYSIS"
title_cell.font = title_font
title_cell.alignment = Alignment(horizontal='center', vertical='center')

# Row 2: Empty spacing row

# Row 3: Column headers with blue background
headers = ["Category / Metric", "Question", "Legend (Possible Values)", 
          "E' Numu Diip\n(Land)", "River Select Foods\n(Fisheries)", 
          "Many Nations\n(Insurance)", "RTZ Artists\n(Arts)"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = blue_header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Set column widths for 7-column layout with Legend
ws.column_dimensions['A'].width = 25.0  # Category
ws.column_dimensions['B'].width = 40.0  # Question
ws.column_dimensions['C'].width = 30.0  # Legend
ws.column_dimensions['D'].width = 15.0  # E' Numu
ws.column_dimensions['E'].width = 15.0  # RSF
ws.column_dimensions['F'].width = 15.0  # Many Nations
ws.column_dimensions['G'].width = 15.0  # RTZ

# Process data starting from row 4
current_row = 4
current_category = ""

for data_row in data[1:]:  # Skip header row
    category = data_row[0].strip()
    question = data_row[1]
    legend = data_row[2]
    values = data_row[3:]
    
    # If new category, add category header row
    if category and category != current_category:
        # Category header row with light blue background
        ws.merge_cells(f'A{current_row}:G{current_row}')
        cat_cell = ws.cell(row=current_row, column=1, value=category.upper())
        cat_cell.font = category_font
        cat_cell.fill = light_blue_fill
        cat_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        current_row += 1
        current_category = category
    
    # Question row
    question_cell = ws.cell(row=current_row, column=2, value=question)
    question_cell.font = question_font
    question_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Legend cell
    legend_cell = ws.cell(row=current_row, column=3, value=legend)
    legend_cell.font = legend_font
    legend_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Data values
    for col, value in enumerate(values, 4):
        if col <= 7:  # Only fill up to column G
            data_cell = ws.cell(row=current_row, column=col, value=value)
            data_cell.font = data_font
            data_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    current_row += 1

# Set row heights
ws.row_dimensions[1].height = 25  # Title row
ws.row_dimensions[3].height = 35  # Header row
for row in range(4, current_row):
    ws.row_dimensions[row].height = 20

# Freeze panes at header row
ws.freeze_panes = 'A4'

# Save the beautiful Excel file with Legend
wb.save('claude-sonnet-4-max_with_legend_v1.6.0.xlsx')

print('✅ Created beautiful Excel with Legend column!')
print('📊 claude-sonnet-4-max_with_legend_v1.6.0.xlsx')
print('📄 claude-sonnet-4-max_with_legend_v1.6.0.csv')
print('')
print('🎨 Beautiful formatting with Legend:')
print('  • Blue header row with white text')
print('  • Light blue category sections')
print('  • Legend column with italicized explanations')
print('  • Professional fonts and spacing')
print('  • Executive title merged across top')
print('  • 7-column layout (Category, Question, Legend, + 4 cooperatives)')
print('  • ALL answers now make logical sense!')

