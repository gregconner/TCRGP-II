import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import csv

# Read the TRULY fixed data
with open('claude-sonnet-4-max_all_fixed_v1.7.0.csv', 'r', encoding='utf-8') as f:
    reader = csv.reader(f)
    data = list(reader)

# Create beautiful Excel with all sensible answers
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Indigenous Cooperative Comparison"

# Beautiful formatting
title_font = Font(name='Calibri', size=16, bold=True)
header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
category_font = Font(name='Calibri', size=11, bold=True)
question_font = Font(name='Calibri', size=10, bold=False)
legend_font = Font(name='Calibri', size=9, bold=False, italic=True)
data_font = Font(name='Calibri', size=10, bold=False)

blue_header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
light_blue_fill = PatternFill(start_color='E8F1FF', end_color='E8F1FF', fill_type='solid')

# Title row
ws.merge_cells('A1:G1')
title_cell = ws['A1']
title_cell.value = "INDIGENOUS COOPERATIVE SYSTEMS - COMPARATIVE ANALYSIS"
title_cell.font = title_font
title_cell.alignment = Alignment(horizontal='center', vertical='center')

# Headers
headers = ["Category", "Question", "Legend", "E' Numu Diip", "River Select Foods", "Many Nations", "RTZ Artists"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = blue_header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Column widths
widths = [25, 40, 30, 15, 15, 15, 15]
for col, width in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = width

# Data rows
current_row = 4
current_category = ""

for data_row in data[1:]:
    category = data_row[0].strip()
    question = data_row[1]
    legend = data_row[2]
    values = data_row[3:]
    
    # Category header
    if category and category != current_category:
        ws.merge_cells(f'A{current_row}:G{current_row}')
        cat_cell = ws.cell(row=current_row, column=1, value=category.upper())
        cat_cell.font = category_font
        cat_cell.fill = light_blue_fill
        cat_cell.alignment = Alignment(horizontal='left', vertical='center')
        current_row += 1
        current_category = category
    
    # Question and data
    ws.cell(row=current_row, column=2, value=question).font = question_font
    ws.cell(row=current_row, column=3, value=legend).font = legend_font
    
    for col, value in enumerate(values, 4):
        if col <= 7:
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.font = data_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    current_row += 1

# Row heights
ws.row_dimensions[1].height = 25
ws.row_dimensions[3].height = 35
for row in range(4, current_row):
    ws.row_dimensions[row].height = 20

ws.freeze_panes = 'A4'

wb.save('claude-sonnet-4-max_all_fixed_v1.7.0.xlsx')

print('✅ Created final Excel with ALL SENSIBLE ANSWERS!')
print('📊 claude-sonnet-4-max_all_fixed_v1.7.0.xlsx')
print('📄 claude-sonnet-4-max_all_fixed_v1.7.0.csv')

