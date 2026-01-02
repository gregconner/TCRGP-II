import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import csv

# Read our clean data (without the Legend column)
with open('claude-sonnet-4-max_clean_readable_matrix_v1.3.0.csv', 'r', encoding='utf-8') as f:
    reader = csv.reader(f)
    data = list(reader)

# Remove the Legend column (column 3) to match previous format
clean_data = []
for row in data:
    new_row = [row[0], row[1]] + row[3:]  # Skip column 2 (Legend)
    clean_data.append(new_row)

# Create new workbook matching the previous beautiful format
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Indigenous Cooperative Comparison"

# Define colors and fonts exactly like the previous version
title_font = Font(name='Calibri', size=16, bold=True)
header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
category_font = Font(name='Calibri', size=11, bold=True)
question_font = Font(name='Calibri', size=10, bold=False)
data_font = Font(name='Calibri', size=10, bold=False)

# Colors from the previous version
blue_header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
light_blue_fill = PatternFill(start_color='E8F1FF', end_color='E8F1FF', fill_type='solid')

# Row 1: Main title (merged across all columns)
ws.merge_cells('A1:F1')
title_cell = ws['A1']
title_cell.value = "INDIGENOUS COOPERATIVE SYSTEMS - COMPARATIVE ANALYSIS"
title_cell.font = title_font
title_cell.alignment = Alignment(horizontal='center', vertical='center')

# Row 2: Empty spacing row

# Row 3: Column headers with blue background
headers = ["Category / Metric", "E' Numu Diip\n(Land Restoration)", "River Select Foods\n(Fisheries)", 
          "Many Nations\n(Insurance)", "RTZ Artists\n(Arts)", ""]
for col, header in enumerate(headers, 1):
    if col <= 5:  # Only first 5 columns get headers
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = blue_header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Set column widths exactly like previous version
ws.column_dimensions['A'].width = 45.0
ws.column_dimensions['B'].width = 18.0
ws.column_dimensions['C'].width = 18.0
ws.column_dimensions['D'].width = 18.0
ws.column_dimensions['E'].width = 18.0
ws.column_dimensions['F'].width = 13.0

# Process data starting from row 4
current_row = 4
current_category = ""

for data_row in clean_data[1:]:  # Skip header row
    category = data_row[0].strip()
    question = data_row[1]
    values = data_row[2:]
    
    # If new category, add category header row
    if category and category != current_category:
        # Category header row with light blue background
        cat_cell = ws.cell(row=current_row, column=1, value=category.upper())
        cat_cell.font = category_font
        cat_cell.fill = light_blue_fill
        cat_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        current_row += 1
        current_category = category
    
    # Question row
    question_cell = ws.cell(row=current_row, column=1, value=question)
    question_cell.font = question_font
    question_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Data values
    for col, value in enumerate(values, 2):
        if col <= 6:  # Only fill up to column F
            data_cell = ws.cell(row=current_row, column=col, value=value)
            data_cell.font = data_font
            data_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    current_row += 1

# Set row heights
ws.row_dimensions[1].height = 25  # Title row
ws.row_dimensions[3].height = 30  # Header row
for row in range(4, current_row):
    ws.row_dimensions[row].height = 18

# Freeze panes at header row
ws.freeze_panes = 'A4'

# Save both Excel and CSV in previous format
wb.save('claude-sonnet-4-max_complete_clean_matrix_v1.3.0.xlsx')

# Also save CSV in previous format (no Legend column)
with open('claude-sonnet-4-max_complete_clean_matrix_v1.3.0.csv', 'w', newline='', encoding='utf-8') as f:
    writer = csv.writer(f)
    for row in clean_data:
        writer.writerow(row)

print('✅ Recreated with previous beautiful formatting!')
print('📊 claude-sonnet-4-max_complete_clean_matrix_v1.3.0.xlsx')
print('📄 claude-sonnet-4-max_complete_clean_matrix_v1.3.0.csv')
print('')
print('🎨 Previous formatting restored:')
print('  • Blue header row with white text')
print('  • Light blue category sections')
print('  • No Legend column (clean 6-column layout)')
print('  • Proper fonts, sizes, and spacing')
print('  • Executive title merged across top')
print('  • All sensible answers maintained')

